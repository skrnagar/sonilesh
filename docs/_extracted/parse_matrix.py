import json
from pathlib import Path

import openpyxl

xlsx = Path(r"C:\Users\hp\Downloads\SafetyApp-Permission-Matrix...xlsx")
wb = openpyxl.load_workbook(xlsx, data_only=True)
out_dir = Path(__file__).parent
result = {"sheets": list(wb.sheetnames)}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    result[sheet_name] = rows
    print(f"Sheet: {sheet_name}, rows={len(rows)}, cols={max(len(r) for r in rows) if rows else 0}")

(out_dir / "permission_matrix.json").write_text(
    json.dumps(result, indent=2), encoding="utf-8"
)
